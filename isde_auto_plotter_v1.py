import os
import re
import shutil

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.utils import get_column_letter


REQUIRED_COLUMNS = {
    "vdd", "input", "ion", "freq", "actual_freq",
    "sr_num", "cs", "upper", "lower"
}


def norm(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def row_is_blank(values):
    return all(v is None or str(v).strip() == "" for v in values)


def extract_let_from_ion(ion_value):
    if pd.isna(ion_value):
        return None
    ion_str = str(ion_value).strip()
    if "-" not in ion_str:
        return None
    try:
        return float(ion_str.split("-")[-1])
    except ValueError:
        return None


def prompt_choice(prompt_text, valid_choices):
    valid_lower = [v.lower() for v in valid_choices]
    while True:
        value = input(prompt_text).strip().lower()
        if value in valid_lower:
            return value
        print(f"Please enter one of: {', '.join(valid_choices)}")


def prompt_float_or_all(prompt_text):
    while True:
        raw = input(prompt_text).strip().lower()
        if raw == "all":
            return "all"
        try:
            return float(raw)
        except ValueError:
            print("Please enter a valid number or 'all'.")


def prompt_input_value(prompt_text):
    raw = input(prompt_text).strip()
    if raw.lower() == "all":
        return "all"
    return raw


def prompt_sheet_choice(sheet_names):
    print("\nAvailable sheets:")
    for i, name in enumerate(sheet_names):
        print(f"{i}: {name}")

    while True:
        raw = input("Choose sheet number: ").strip()
        try:
            idx = int(raw)
            if 0 <= idx < len(sheet_names):
                return sheet_names[idx]
            print("Invalid sheet number.")
        except ValueError:
            print("Please enter a valid integer.")


def parse_shift_register_token(token):
    """
    Accepts:
      A-S5
      Z-S12
      S8
      8
    """
    token = token.strip().upper()
    if not token:
        return None

    m = re.fullmatch(r"([AZ])\-S(\d+)", token)
    if m:
        return {"prefix": m.group(1), "number": int(m.group(2)), "raw": token}

    m = re.fullmatch(r"S(\d+)", token)
    if m:
        return {"prefix": None, "number": int(m.group(1)), "raw": token}

    m = re.fullmatch(r"(\d+)", token)
    if m:
        return {"prefix": None, "number": int(m.group(1)), "raw": token}

    return None


def prompt_shift_registers(prompt_text):
    while True:
        raw = input(prompt_text).strip()
        if raw.lower() == "all":
            return "all"

        parts = [x.strip() for x in raw.split(",") if x.strip()]
        parsed = [parse_shift_register_token(x) for x in parts]

        if parts and all(p is not None for p in parsed):
            return parsed

        print("Enter comma-separated shift registers like: A-S5, Z-S10, S3, or type 'all'")


def find_tables_in_sheet(ws):
    tables = []
    max_row = ws.max_row
    max_col = ws.max_column
    scanned_header_rows = set()

    for r in range(1, max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        normalized = [norm(v) for v in row_vals]

        present = set(v for v in normalized if v)
        if not REQUIRED_COLUMNS.issubset(present):
            continue

        if r in scanned_header_rows:
            continue

        header_positions = {}
        for c_idx, cell_val in enumerate(normalized, start=1):
            if cell_val in REQUIRED_COLUMNS and cell_val not in header_positions:
                header_positions[cell_val] = c_idx

        if not REQUIRED_COLUMNS.issubset(set(header_positions.keys())):
            continue

        scanned_header_rows.add(r)

        data_rows = []
        blank_streak = 0

        for rr in range(r + 1, max_row + 1):
            row_dict = {}
            for col_name, col_idx in header_positions.items():
                row_dict[col_name] = ws.cell(rr, col_idx).value

            if row_is_blank(list(row_dict.values())):
                blank_streak += 1
                if blank_streak >= 2:
                    break
                continue
            else:
                blank_streak = 0

            data_rows.append(row_dict)

        if data_rows:
            df = pd.DataFrame(data_rows)
            df["source_sheet"] = ws.title
            df["header_row"] = r
            tables.append(df)

    return tables


def clean_table(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    rename_map = {}
    for c in df.columns:
        lc = c.strip().lower()
        if lc in REQUIRED_COLUMNS:
            rename_map[c] = lc
    df = df.rename(columns=rename_map)

    if "input" in df.columns:
        df["input"] = df["input"].astype(str).str.strip()

    numeric_cols = ["vdd", "freq", "actual_freq", "cs", "upper", "lower"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["sr_num_raw"] = df["sr_num"].astype(str).str.strip()

    extracted_sr_num = df["sr_num_raw"].str.extract(r"(\d+)", expand=False)
    df["sr_num_numeric"] = pd.to_numeric(extracted_sr_num, errors="coerce")

    extracted_prefix = df["sr_num_raw"].str.extract(r"\b([AZ])\-S\d+\b", expand=False)
    df["sr_prefix"] = extracted_prefix.str.upper()

    if "ion" in df.columns:
        df["LET"] = df["ion"].apply(extract_let_from_ion)
    else:
        df["LET"] = None

    return df


def float_matches(series, value, tol=1e-9):
    return series.notna() & ((series - value).abs() < tol)


def filter_table(df, x_axis, input_val, selected_srs, vdd_val=None, let_val=None, freq_val=None):
    filtered = df.copy()

    if str(input_val).strip().lower() != "all":
        filtered = filtered[filtered["input"] == str(input_val).strip()]

    if selected_srs != "all":
        selected_numbers = sorted(set(sr["number"] for sr in selected_srs))
        filtered = filtered[filtered["sr_num_numeric"].isin(selected_numbers)]

        has_prefix_data = filtered["sr_prefix"].notna().any()
        if has_prefix_data:
            allowed_pairs = {(sr["prefix"], sr["number"]) for sr in selected_srs if sr["prefix"] is not None}
            no_prefix_requests = {sr["number"] for sr in selected_srs if sr["prefix"] is None}

            keep_mask = []
            for _, row in filtered.iterrows():
                sr_num = int(row["sr_num_numeric"]) if pd.notna(row["sr_num_numeric"]) else None
                pair = (row["sr_prefix"], sr_num)

                keep = False
                if sr_num in no_prefix_requests:
                    keep = True
                if pair in allowed_pairs:
                    keep = True

                keep_mask.append(keep)

            filtered = filtered[pd.Series(keep_mask, index=filtered.index)]

    if x_axis != "vdd" and vdd_val != "all":
        filtered = filtered[float_matches(filtered["vdd"], vdd_val)]

    if x_axis != "let" and let_val != "all":
        filtered = filtered[float_matches(filtered["LET"], let_val)]

    if x_axis != "frq" and freq_val != "all":
        filtered = filtered[float_matches(filtered["freq"], freq_val)]

    return filtered


def format_vdd_title(vdd_val):
    if vdd_val is None or vdd_val == "all":
        return None
    if vdd_val < 10:
        return f"{int(round(vdd_val * 1000))} mV"
    return f"{int(round(vdd_val))} mV"


def format_freq_title(freq_val):
    if freq_val is None or freq_val == "all":
        return None
    if float(freq_val).is_integer():
        return f"{int(freq_val)} MHz"
    return f"{freq_val} MHz"


def format_let_title(let_val):
    if let_val is None or let_val == "all":
        return None
    if float(let_val).is_integer():
        return f"LET {int(let_val)}"
    return f"LET {let_val}"


def build_plot_title(x_axis, input_val, vdd_val=None, let_val=None, freq_val=None):
    title_parts = []

    if x_axis != "vdd" and vdd_val not in [None, "all"]:
        title_parts.append(format_vdd_title(vdd_val))

    if x_axis != "frq" and freq_val not in [None, "all"]:
        title_parts.append(format_freq_title(freq_val))

    if x_axis != "let" and let_val not in [None, "all"]:
        title_parts.append(format_let_title(let_val))

    if str(input_val).strip().lower() != "all":
        title_parts.append(f"Input {input_val}")
    else:
        title_parts.append("All Inputs")

    return " ".join(title_parts)


def build_series_label_from_group(sr_df, requested_srs, sr_num, x_axis):
    prefixes = sorted(set(x for x in sr_df["sr_prefix"].dropna().unique()))
    if len(prefixes) == 1:
        base = f"{prefixes[0]}-S{int(sr_num)}"
    elif len(prefixes) > 1:
        base = f"S{int(sr_num)}"
    else:
        if requested_srs != "all":
            matching_requests = [sr for sr in requested_srs if sr["number"] == int(sr_num)]
            if len(matching_requests) == 1 and matching_requests[0]["prefix"] is not None:
                base = matching_requests[0]["raw"]
            else:
                base = f"S{int(sr_num)}"
        else:
            base = f"S{int(sr_num)}"

    extras = []

    if x_axis != "vdd" and sr_df["vdd"].notna().any():
        extras.append(format_vdd_title(sr_df["vdd"].iloc[0]))

    if x_axis != "let" and sr_df["LET"].notna().any():
        extras.append(format_let_title(sr_df["LET"].iloc[0]))

    if x_axis != "frq" and sr_df["freq"].notna().any():
        extras.append(format_freq_title(sr_df["freq"].iloc[0]))

    if sr_df["input"].notna().any():
        input_value = str(sr_df["input"].iloc[0]).strip()
        extras.append(f"Input {input_value}")

    if extras:
        return f"{base} ({', '.join(extras)})"
    return base


def write_helper_data_for_chart(ws, combined, selected_srs, helper_start_col, x_axis):
    current_col = helper_start_col
    series_meta = []

    group_cols = ["sr_num_numeric", "sr_prefix"]

    if x_axis != "vdd":
        group_cols.append("vdd")
    if x_axis != "let":
        group_cols.append("LET")
    if x_axis != "frq":
        group_cols.append("freq")

    group_cols.append("input")

    grouped = combined.groupby(group_cols, dropna=False)

    for _, sr_df in grouped:
        sr_df = sr_df.sort_values("x").copy()
        if sr_df.empty:
            continue

        sr_num = sr_df["sr_num_numeric"].iloc[0]
        label = build_series_label_from_group(sr_df, selected_srs, sr_num, x_axis)

        x_col = current_col
        y_col = current_col + 1
        plus_col = current_col + 2
        minus_col = current_col + 3

        ws.cell(row=1, column=x_col, value=f"{label}_x")
        ws.cell(row=1, column=y_col, value=f"{label}_y")
        ws.cell(row=1, column=plus_col, value=f"{label}_plus")
        ws.cell(row=1, column=minus_col, value=f"{label}_minus")

        start_row = 2
        for i, (_, row) in enumerate(sr_df.iterrows(), start=start_row):
            ws.cell(row=i, column=x_col, value=float(row["x"]))
            ws.cell(row=i, column=y_col, value=float(row["y"]))
            ws.cell(row=i, column=plus_col, value=float(row["upper"]))
            ws.cell(row=i, column=minus_col, value=float(row["lower"]))

        end_row = start_row + len(sr_df) - 1

        series_meta.append({
            "label": label,
            "x_col": x_col,
            "y_col": y_col,
            "plus_col": plus_col,
            "minus_col": minus_col,
            "start_row": start_row,
            "end_row": end_row,
        })

        for col_idx in [x_col, y_col, plus_col, minus_col]:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

        current_col += 5

    return series_meta


def add_native_excel_chart(ws, sheet_name, series_meta, x_axis, scale, chart_title, chart_anchor):
    chart = ScatterChart()
    chart.title = chart_title
    chart.style = 2
    chart.height = 14
    chart.width = 24

    if x_axis == "vdd":
        chart.x_axis.title = "VDD"
    elif x_axis == "let":
        chart.x_axis.title = "LET"
    else:
        chart.x_axis.title = "Frequency"

    chart.y_axis.title = "Cross Section"
    chart.legend.position = "r"

    if scale == "log":
        chart.y_axis.scaling.logBase = 10
        if x_axis == "frq":
            chart.x_axis.scaling.logBase = 10

    for meta in series_meta:
        xref = Reference(
            ws,
            min_col=meta["x_col"],
            min_row=meta["start_row"],
            max_row=meta["end_row"]
        )
        yref = Reference(
            ws,
            min_col=meta["y_col"],
            min_row=meta["start_row"],
            max_row=meta["end_row"]
        )

        series = Series(yref, xref, title=meta["label"])
        series.marker.symbol = "circle"
        series.graphicalProperties.line.width = 19050

        plus_range = f"'{sheet_name}'!${get_column_letter(meta['plus_col'])}${meta['start_row']}:${get_column_letter(meta['plus_col'])}${meta['end_row']}"
        minus_range = f"'{sheet_name}'!${get_column_letter(meta['minus_col'])}${meta['start_row']}:${get_column_letter(meta['minus_col'])}${meta['end_row']}"

        err_bars = ErrorBars()
        err_bars.errDir = "y"
        err_bars.errBarType = "both"
        err_bars.noEndCap = False
        err_bars.plus = NumDataSource(numRef=NumRef(f=plus_range))
        err_bars.minus = NumDataSource(numRef=NumRef(f=minus_range))
        series.errBars = err_bars

        chart.series.append(series)

    ws.add_chart(chart, chart_anchor)


def main():
    print("=== ISDE Automatic Plotter (Native Excel Chart) ===")

    file_path = input("Excel file path: ").strip()

    if not os.path.exists(file_path):
        print("File not found.")
        return

    base, ext = os.path.splitext(file_path)
    new_file = base + "_with_graph.xlsx"
    shutil.copy(file_path, new_file)

    wb_data = load_workbook(new_file, data_only=True)

    selected_sheet_name = prompt_sheet_choice(wb_data.sheetnames)
    ws_data = wb_data[selected_sheet_name]

    print("\nEnter plot settings:\n")

    x_axis = prompt_choice("Choose x-axis (vdd, let, frq): ", ["vdd", "let", "frq"])
    input_val = prompt_input_value("Input value (or type 'all'): ")
    selected_srs = prompt_shift_registers(
        "Shift registers (comma separated, e.g. A-S5, Z-S10, or type 'all'): "
    )
    scale = prompt_choice("Scale (linear/log): ", ["linear", "log"])

    vdd_val = None
    let_val = None
    freq_val = None

    if x_axis != "vdd":
        vdd_val = prompt_float_or_all("Specify VDD (or type 'all'): ")

    if x_axis != "let":
        let_val = prompt_float_or_all("Specify LET (or type 'all'): ")

    if x_axis != "frq":
        freq_val = prompt_float_or_all("Specify freq (or type 'all'): ")

    print(f"\nScanning sheet '{selected_sheet_name}' for matching embedded tables...")

    all_tables = find_tables_in_sheet(ws_data)

    if not all_tables:
        print("No tables with the required columns were found in that sheet.")
        return

    matched_frames = []

    for raw_df in all_tables:
        try:
            df = clean_table(raw_df)

            filtered = filter_table(
                df=df,
                x_axis=x_axis,
                input_val=input_val,
                selected_srs=selected_srs,
                vdd_val=vdd_val,
                let_val=let_val,
                freq_val=freq_val
            )

            if not filtered.empty:
                matched_frames.append(filtered)

        except Exception:
            continue

    if not matched_frames:
        print("No matching data found in the selected sheet.")
        return

    combined = pd.concat(matched_frames, ignore_index=True)

    if x_axis == "vdd":
        combined["x"] = combined["vdd"]
    elif x_axis == "let":
        combined["x"] = combined["LET"]
    else:
        combined["x"] = combined["actual_freq"]

    combined["y"] = combined["cs"]
    combined = combined.dropna(subset=["x", "y", "lower", "upper", "sr_num_numeric", "input"])

    if combined.empty:
        print("Matching data was found, but plotting columns were invalid after cleaning.")
        return

    combined = combined.sort_values(["sr_num_numeric", "input", "x"])

    plot_title = build_plot_title(
        x_axis=x_axis,
        input_val=input_val,
        vdd_val=vdd_val,
        let_val=let_val,
        freq_val=freq_val
    )

    used_locations = combined[["source_sheet", "header_row"]].drop_duplicates()
    print("\nMatched table locations:")
    for _, row in used_locations.iterrows():
        print(f"- Sheet: {row['source_sheet']}, header row: {int(row['header_row'])}")

    wb_write = load_workbook(new_file)
    ws_write = wb_write[selected_sheet_name]

    original_max_col = ws_write.max_column
    helper_start_col = original_max_col + 20
    chart_anchor_col = original_max_col + 2
    chart_anchor = f"{get_column_letter(chart_anchor_col)}2"

    series_meta = write_helper_data_for_chart(
        ws=ws_write,
        combined=combined,
        selected_srs=selected_srs,
        helper_start_col=helper_start_col,
        x_axis=x_axis
    )

    if not series_meta:
        print("No plottable data found.")
        return

    add_native_excel_chart(
        ws=ws_write,
        sheet_name=selected_sheet_name,
        series_meta=series_meta,
        x_axis=x_axis,
        scale=scale,
        chart_title=plot_title,
        chart_anchor=chart_anchor
    )

    wb_write.save(new_file)

    print(f"\nDone. Native Excel chart inserted into sheet '{selected_sheet_name}' in {new_file}")


if __name__ == "__main__":
    main()
